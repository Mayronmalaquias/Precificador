from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_MONEY_FORMAT = 'R$ #,##0.00'


def _write_table(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 45)
        ws.column_dimensions[get_column_letter(column[0].column)].width = width


def gerar_xlsx_vgc_captacao(relatorio):
    """Gera o relatório de VGC cruzado com foco PP/AC da base de imóveis."""
    wb = Workbook()
    resumo_ws = wb.active
    resumo_ws.title = "Resumo Foco"

    resumo_rows = [
        [
            item["classificacao_foco"],
            item["qtd_negocios"],
            item["vgc_total"],
            item["vgv_total"],
        ]
        for item in relatorio["resumo"]
    ]
    _write_table(
        resumo_ws,
        ["Classificação", "Negócios", "VGC Total", "VGV Total"],
        resumo_rows,
    )
    for row in resumo_ws.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row:
            cell.number_format = _MONEY_FORMAT

    ranking_ws = wb.create_sheet("Ranking por Foco")
    ranking_rows = [
        [
            item["classificacao_foco"], item["corretor"], item["qtd_negocios"],
            item["vgc_corretor"], item["vgv_imoveis"],
        ]
        for item in relatorio["ranking"]
    ]
    _write_table(
        ranking_ws,
        ["Classificação", "Corretor", "Negócios", "VGC do Corretor", "VGV dos Imóveis"],
        ranking_rows,
    )
    for row in ranking_ws.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = _MONEY_FORMAT

    detalhe_ws = wb.create_sheet("Negócios por Foco")
    detalhe_rows = [
        [
            item["data_contrato"],
            item["id_contrato"],
            item["contrato"],
            item["codigo_imovel"],
            item["empreendimento"],
            item["classificacao_foco"],
            item["corretor"],
            item["papel"],
            item["captadores"],
            item["vendedores"],
            item["valor_negocio"],
            item["vgc_61_contrato"],
            item["vgc_do_corretor"],
        ]
        for item in relatorio["detalhes"]
    ]
    _write_table(
        detalhe_ws,
        [
            "Data", "ID Contrato", "Contrato", "Código do Imóvel", "Empreendimento",
            "Classificação", "Corretor", "Papel", "Captadores", "Vendedores",
            "Valor do Negócio", "VGC 61 do Contrato", "VGC do Corretor",
        ],
        detalhe_rows,
    )
    for row in detalhe_ws.iter_rows(min_row=2, min_col=11, max_col=13):
        for cell in row:
            cell.number_format = _MONEY_FORMAT

    info_ws = wb.create_sheet("Critérios")
    criterios = [
        ["Relatório", "VGC das captações por foco do imóvel"],
        ["Período inicial", relatorio.get("start") or "Todos"],
        ["Período final", relatorio.get("end") or "Todos"],
        ["Fator 6% aplicado", "Sim" if relatorio.get("apply_factor") else "Não"],
        ["Origem do foco", "contratos.Codigo_Imovel → imoveis_legado.codigo (Foco PP / Foco AC)"],
        ["Regra VGC", "O VGC segue a mesma divisão entre os participantes usada pelo ranking."],
        ["Total VGC", relatorio.get("total_vgc", 0)],
    ]
    for row in criterios:
        info_ws.append(row)
    info_ws.column_dimensions["A"].width = 24
    info_ws.column_dimensions["B"].width = 100
    info_ws["B7"].number_format = _MONEY_FORMAT
    info_ws["A1"].font = Font(bold=True)
    info_ws["B1"].font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
